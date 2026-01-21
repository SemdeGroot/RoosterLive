# core/utils/beat/uren.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Dict, Optional, Tuple, Set, List, Any

from dateutil.relativedelta import relativedelta
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db.models import Count, Q
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import Dagdeel, Shift, UrenMaand, UrenRegel


@dataclass(frozen=True)
class UrenExportResult:
    month: date
    xlsx_storage_path: Optional[str]
    filename: Optional[str]
    row_count: int


def _month_first(d: date) -> date:
    return d.replace(day=1)


def _prev_month_first(today: date) -> date:
    return _month_first(today + relativedelta(months=-1))


def _next_month_first(month_first: date) -> date:
    return _month_first(month_first + relativedelta(months=1))


def _dutch_name(user) -> str:
    """
    Zelfde gedrag als jouw templatetag: eerste en laatste woord met hoofdletter.
    """
    if not user:
        return "-"

    full_name = f"{(user.first_name or '').strip()} {(user.last_name or '').strip()}".strip()
    if not full_name:
        return (getattr(user, "username", "") or "").strip() or f"User {user.pk}"

    parts = full_name.split()
    if not parts:
        return full_name

    parts[0] = parts[0].capitalize()
    if len(parts) > 1:
        parts[-1] = parts[-1].capitalize()
    return " ".join(parts)


def _autosize_columns(ws, min_width: int = 12, max_width: int = 45) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def _dagdeel_minutes(d: Dagdeel) -> int:
    def to_minutes(t):
        return t.hour * 60 + t.minute

    s = to_minutes(d.start_time)
    e = to_minutes(d.end_time)
    if e == s:
        return 0
    if e > s:
        return e - s
    return (1440 - s) + e


def _dagdeel_hours_1_decimal(d: Dagdeel) -> Decimal:
    mins = _dagdeel_minutes(d)
    if mins <= 0:
        return Decimal("0.0")
    hrs = Decimal(mins) / Decimal(60)
    return hrs.quantize(Decimal("0.1"))


def _shift_period_to_dagdeel_code(period: str) -> str:
    # jouw mapping
    if period == "morning":
        return Dagdeel.CODE_MORNING
    if period == "afternoon":
        return Dagdeel.CODE_AFTERNOON
    return Dagdeel.CODE_PRE_EVENING


def _excel_apply_common_header_style(ws, header_row: int, ncols: int) -> None:
    """
    Licht headervlak + zwart font (werkt ook als fills slecht renderen in LibreOffice).
    """
    header_fill = PatternFill("solid", fgColor="E5E7EB")  # lichtgrijs
    header_font = Font(bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align


def _style_status_cell(cell, *, text: str, kind: str) -> None:
    """
    kind:
      - "ok"  -> groen
      - "bad" -> rood
      - "muted" -> grijs
    """
    if kind == "ok":
        cell.font = Font(bold=True, color="065F46")  # donker groen
    elif kind == "bad":
        cell.font = Font(bold=True, color="991B1B")  # donker rood
    else:
        cell.font = Font(bold=True, color="374151")  # donker grijs
    cell.value = text
    cell.alignment = Alignment(horizontal="center", vertical="center")


def export_uren_month_to_storage(today: Optional[date] = None) -> UrenExportResult:
    """
    Exporteert vorige kalendermaand (1e t/m 1e volgende maand).

    TAB 1 (Overzicht):
    - Totaal doorgegeven uren (gewogen): sum(actual_hours * allowance_multiplier)
    - Totaal geschatte uren (gewogen): shifts per dagdeel * dagdeel_duur * allowance_multiplier(dagdeel)
    - Verschil (uren): doorgegeven - geschat
    - Verschil (%): (doorgegeven - geschat) / geschat * 100
    - Kilometers: UrenMaand.kilometers (0 als ontbreekt)
    - Sortering: grootste absolute verschil in UREN bovenaan (abs(verschil_uren) desc)

    TAB 2 (Details):
    Werknemer | Datum | Shift | Dagdeel | Theoretisch (gewogen) uren | Doorgegeven (gewogen) uren | Verschil (uren) | Verschil (%)

    Regels in Details:
    1) Voor geplande shifts:
       - Shift kolom: "Ingepland" (groen)
       - Theoretisch gevuld (dagdeel_duur * multiplier)
       - Doorgegeven is gewogen uit UrenRegel (of "Geen uren" rood)
    2) Voor alle doorgegeven uren ZONDER bijbehorende shift:
       - Shift kolom: "Geen shift" (rood)
       - Theoretisch leeg
       - Verschil(uren)=doorgegeven
       - Verschil(%) leeg
    Details wordt gesorteerd op: werknemer -> datum -> dagdeel.sort_order
    """
    if today is None:
        today = timezone.localdate()

    month_first = _prev_month_first(today)
    next_month = _next_month_first(month_first)

    # Dagdelen + mapping
    dagdelen = list(Dagdeel.objects.all())
    dagdeel_by_code: Dict[str, Dagdeel] = {d.code: d for d in dagdelen}

    # Planning dagdelen (voor shift schatting)
    planning_codes = [
        Dagdeel.CODE_MORNING,
        Dagdeel.CODE_AFTERNOON,
        Dagdeel.CODE_PRE_EVENING,
    ]
    for c in planning_codes:
        if c not in dagdeel_by_code:
            return UrenExportResult(month=month_first, xlsx_storage_path=None, filename=None, row_count=0)

    # durations + multipliers voor geplande shifts
    d_m = dagdeel_by_code[Dagdeel.CODE_MORNING]
    d_a = dagdeel_by_code[Dagdeel.CODE_AFTERNOON]
    d_e = dagdeel_by_code[Dagdeel.CODE_PRE_EVENING]

    dur_m = _dagdeel_hours_1_decimal(d_m)
    dur_a = _dagdeel_hours_1_decimal(d_a)
    dur_e = _dagdeel_hours_1_decimal(d_e)

    mult_m = Decimal(d_m.allowance_pct) / Decimal("100.0")
    mult_a = Decimal(d_a.allowance_pct) / Decimal("100.0")
    mult_e = Decimal(d_e.allowance_pct) / Decimal("100.0")

    # --- Userset: iedereen met uren óf shifts in deze maand ---
    user_ids_from_hours = set(
        UrenRegel.objects.filter(month=month_first, actual_hours__isnull=False)
        .values_list("user_id", flat=True)
        .distinct()
    )
    user_ids_from_shifts = set(
        Shift.objects.filter(date__gte=month_first, date__lt=next_month)
        .values_list("user_id", flat=True)
        .distinct()
    )
    user_ids = sorted(user_ids_from_hours | user_ids_from_shifts)
    if not user_ids:
        return UrenExportResult(month=month_first, xlsx_storage_path=None, filename=None, row_count=0)

    # --- Kilometers per user ---
    km_by_user: Dict[int, int] = {
        r["user_id"]: int(r["kilometers"] or 0)
        for r in UrenMaand.objects.filter(month=month_first, user_id__in=user_ids).values("user_id", "kilometers")
    }

    # --- Urenregels (voor totals + details) ---
    regels = list(
        UrenRegel.objects.select_related("dagdeel", "user")
        .filter(month=month_first, user_id__in=user_ids, actual_hours__isnull=False)
        .order_by("user_id", "date", "dagdeel__sort_order", "id")
    )

    # reported weighted totals
    reported_by_user: Dict[int, Decimal] = {uid: Decimal("0.0") for uid in user_ids}
    user_obj_by_id: Dict[int, Any] = {}

    # reported by (uid, date, dagdeel_code) for details matching
    reported_by_key: Dict[Tuple[int, date, str], Decimal] = {}

    for r in regels:
        user_obj_by_id.setdefault(r.user_id, r.user)
        mult = Decimal(r.dagdeel.allowance_pct) / Decimal("100.0")
        weighted = (r.actual_hours or Decimal("0.0")) * mult

        reported_by_user[r.user_id] += weighted
        key = (r.user_id, r.date, r.dagdeel.code)
        reported_by_key[key] = reported_by_key.get(key, Decimal("0.0")) + weighted

    # Zorg dat we ook users hebben die alleen shifts hadden (maar geen urenregels)
    if len(user_obj_by_id) < len(user_ids):
        missing = [uid for uid in user_ids if uid not in user_obj_by_id]
        if missing:
            for s in (
                Shift.objects.select_related("user")
                .filter(user_id__in=missing, date__gte=month_first, date__lt=next_month)
                .only("user_id", "user__first_name", "user__last_name", "user__email", "user__username")
            ):
                user_obj_by_id.setdefault(s.user_id, s.user)

    # --- Shifts tellen per user (voor overzicht schatting) ---
    shift_counts = (
        Shift.objects.filter(user_id__in=user_ids, date__gte=month_first, date__lt=next_month)
        .values("user_id")
        .annotate(
            morning_count=Count("id", filter=Q(period="morning")),
            afternoon_count=Count("id", filter=Q(period="afternoon")),
            evening_count=Count("id", filter=Q(period="evening")),
        )
    )
    shift_by_user: Dict[int, Dict[str, int]] = {
        row["user_id"]: {
            "morning": int(row.get("morning_count") or 0),
            "afternoon": int(row.get("afternoon_count") or 0),
            "evening": int(row.get("evening_count") or 0),
        }
        for row in shift_counts
    }

    # --- Build overzicht rows then sort on abs diff hours desc ---
    overzicht_rows: List[Tuple[int, str, Decimal, Decimal, Decimal | None, int]] = []
    for uid in user_ids:
        user = user_obj_by_id.get(uid)
        name = _dutch_name(user) if user else f"User {uid}"

        reported = reported_by_user.get(uid, Decimal("0.0"))

        counts = shift_by_user.get(uid, {"morning": 0, "afternoon": 0, "evening": 0})
        m_cnt = Decimal(counts["morning"])
        a_cnt = Decimal(counts["afternoon"])
        e_cnt = Decimal(counts["evening"])

        estimated = (m_cnt * dur_m * mult_m) + (a_cnt * dur_a * mult_a) + (e_cnt * dur_e * mult_e)

        diff_pct = None
        if estimated and estimated != Decimal("0.0"):
            diff_pct = (reported - estimated) / estimated * Decimal("100.0")

        km = km_by_user.get(uid, 0)

        overzicht_rows.append((uid, name, reported, estimated, diff_pct, km))

    # Sort: grootste absolute verschil in UREN eerst
    # tie-break: abs % (als aanwezig), daarna naam
    overzicht_rows.sort(
        key=lambda t: (
            -(abs(t[2] - t[3])),  # abs(reported-estimated) desc
            0 if t[4] is not None else 1,
            -(abs(t[4]) if t[4] is not None else Decimal("0.0")),
            (t[1] or "").lower(),
        )
    )

    # --- Workbook ---
    wb = Workbook()

    # ========== Sheet 1: Overzicht ==========
    ws = wb.active
    ws.title = f"Overzicht {month_first.strftime('%Y-%m')}"

    headers = [
        "Werknemer",
        "Totaal doorgegeven uren (gewogen)",
        "Totaal geschatte uren (gewogen)",
        "Verschil (uren)",
        "Verschil (%)",
        "Totaal kilometers",
    ]
    ws.append(headers)
    _excel_apply_common_header_style(ws, header_row=1, ncols=len(headers))

    for (_uid, name, reported, estimated, diff_pct, km) in overzicht_rows:
        diff_hours = reported - estimated
        ws.append(
            [
                name,
                float(reported),
                float(estimated),
                float(diff_hours),
                (float(diff_pct) if diff_pct is not None else None),
                int(km),
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].alignment = Alignment(horizontal="left", vertical="center")

        row[1].number_format = "0.0"
        row[2].number_format = "0.0"
        row[3].number_format = "0.0"
        row[4].number_format = "0.0"
        row[5].number_format = "0"

        for cell in row[1:]:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Conditional formatting (verschil %) op kolom E
    red_fill = PatternFill("solid", fgColor="FDE2E2")
    orange_fill = PatternFill("solid", fgColor="FEF3C7")

    ws.conditional_formatting.add(
        f"E2:E{ws.max_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["10"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        f"E2:E{ws.max_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["-10"], fill=red_fill),
    )
    ws.conditional_formatting.add(
        f"E2:E{ws.max_row}",
        CellIsRule(operator="between", formula=["5", "10"], fill=orange_fill),
    )
    ws.conditional_formatting.add(
        f"E2:E{ws.max_row}",
        CellIsRule(operator="between", formula=["-10", "-5"], fill=orange_fill),
    )

    _autosize_columns(ws, min_width=12, max_width=55)

    # ========== Sheet 2: Details ==========
    ws2 = wb.create_sheet(title="Details")

    details_headers = [
        "Werknemer",
        "Datum",
        "Shift",
        "Dagdeel",
        "Theoretisch (gewogen) uren",
        "Doorgegeven (gewogen) uren",
        "Verschil (uren)",
        "Verschil (%)",
    ]
    ws2.append(details_headers)
    _excel_apply_common_header_style(ws2, header_row=1, ncols=len(details_headers))

    # Alle shifts deze maand (voor details)
    shifts = list(
        Shift.objects.select_related("user")
        .filter(user_id__in=user_ids, date__gte=month_first, date__lt=next_month)
        .only("user_id", "date", "period", "user__first_name", "user__last_name", "user__email", "user__username")
        .order_by("user_id", "date", "period")
    )

    # Keys die een geplande shift representeren (voor het detecteren van "doorgegeven zonder shift")
    matched_shift_keys: Set[Tuple[int, date, str]] = set()

    # We bouwen eerst een lijst records, dan sorteren we exact zoals je wilt:
    # werknemer -> datum -> dagdeel.sort_order
    details_rows: List[Dict[str, Any]] = []

    # 1) Shift-rijen: theoretisch gevuld, doorgegeven optioneel
    for s in shifts:
        user_obj_by_id.setdefault(s.user_id, s.user)
        user_obj = user_obj_by_id.get(s.user_id)
        name = _dutch_name(user_obj) if user_obj else f"User {s.user_id}"

        dagdeel_code = _shift_period_to_dagdeel_code(s.period)
        dagdeel = dagdeel_by_code.get(dagdeel_code)
        if not dagdeel:
            continue

        matched_shift_keys.add((s.user_id, s.date, dagdeel_code))

        if dagdeel_code == Dagdeel.CODE_MORNING:
            theoretical = dur_m * mult_m
        elif dagdeel_code == Dagdeel.CODE_AFTERNOON:
            theoretical = dur_a * mult_a
        else:
            theoretical = dur_e * mult_e

        reported = reported_by_key.get((s.user_id, s.date, dagdeel_code))  # gewogen

        if reported is None:
            diff_hours = None
            diff_pct = None
        else:
            diff_hours = reported - theoretical
            diff_pct = (diff_hours / theoretical * Decimal("100.0")) if theoretical != Decimal("0.0") else None

        details_rows.append(
            {
                "uid": s.user_id,
                "name": name,
                "date": s.date,
                "dagdeel_sort": int(dagdeel.sort_order or 0),
                "shift_kind": "planned",  # ingepland
                "shift_text": "Ingepland",
                "dagdeel_label": dagdeel.get_code_display(),
                "theoretical": theoretical,
                "reported": reported,
                "diff_hours": diff_hours,
                "diff_pct": diff_pct,
            }
        )

    # 2) Doorgegeven zonder shift: ALLES wat in reported_by_key zit maar NIET in matched_shift_keys
    for (uid, d, dagdeel_code), reported in reported_by_key.items():
        if (uid, d, dagdeel_code) in matched_shift_keys:
            continue

        dagdeel = dagdeel_by_code.get(dagdeel_code)
        if not dagdeel:
            continue

        user_obj = user_obj_by_id.get(uid)
        name = _dutch_name(user_obj) if user_obj else f"User {uid}"

        details_rows.append(
            {
                "uid": uid,
                "name": name,
                "date": d,
                "dagdeel_sort": int(dagdeel.sort_order or 0),
                "shift_kind": "missing_shift",  # geen shift
                "shift_text": "Geen shift",
                "dagdeel_label": dagdeel.get_code_display(),
                "theoretical": None,          # expliciete wens
                "reported": reported,         # gewogen
                "diff_hours": reported,       # verschil t.o.v. 0 gepland
                "diff_pct": None,             # leeg
            }
        )

    # Sort: werknemer -> datum -> dagdeel.sort_order
    details_rows.sort(key=lambda r: ((r["name"] or "").lower(), r["date"], r["dagdeel_sort"]))

    # Schrijf rows weg + style status/lege velden
    for r in details_rows:
        row_idx = ws2.max_row + 1

        ws2.append(
            [
                r["name"],
                r["date"].strftime("%d-%m-%Y"),
                "",  # shift tekst vullen we gestyled hieronder
                r["dagdeel_label"],
                (float(r["theoretical"]) if r["theoretical"] is not None else None),
                (float(r["reported"]) if r["reported"] is not None else None),
                (float(r["diff_hours"]) if r["diff_hours"] is not None else None),
                (float(r["diff_pct"]) if r["diff_pct"] is not None else None),
            ]
        )

        # Shift cel: "Ingepland" groen, "Geen shift" rood
        shift_cell = ws2.cell(row=row_idx, column=3)
        if r["shift_kind"] == "planned":
            _style_status_cell(shift_cell, text="Ingepland", kind="ok")
        else:
            _style_status_cell(shift_cell, text="Geen shift", kind="bad")

        # Doorgegeven leeg? (alleen bij geplande shifts kan dat)
        if r["shift_kind"] == "planned" and r["reported"] is None:
            rep_cell = ws2.cell(row=row_idx, column=6)
            _style_status_cell(rep_cell, text="Geen uren", kind="bad")

    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:H{ws2.max_row}"

    # Uitlijning + formats
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        row[0].alignment = Alignment(horizontal="left", vertical="center")
        row[1].alignment = Alignment(horizontal="center", vertical="center")  # datum
        # shift col (3) is al gestyled
        row[3].alignment = Alignment(horizontal="center", vertical="center")  # dagdeel

        # Theoretisch / Doorgegeven / Verschil uren
        row[4].number_format = "0.0"
        row[5].number_format = "0.0"
        row[6].number_format = "0.0"
        row[7].number_format = "0.0"

        for cell in row[4:]:
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Conditional formatting op Verschil (%) (kolom H)
    ws2.conditional_formatting.add(
        f"H2:H{ws2.max_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["10"], fill=red_fill),
    )
    ws2.conditional_formatting.add(
        f"H2:H{ws2.max_row}",
        CellIsRule(operator="lessThanOrEqual", formula=["-10"], fill=red_fill),
    )
    ws2.conditional_formatting.add(
        f"H2:H{ws2.max_row}",
        CellIsRule(operator="between", formula=["5", "10"], fill=orange_fill),
    )
    ws2.conditional_formatting.add(
        f"H2:H{ws2.max_row}",
        CellIsRule(operator="between", formula=["-10", "-5"], fill=orange_fill),
    )

    _autosize_columns(ws2, min_width=12, max_width=65)

    # Save to storage
    bio = BytesIO()
    wb.save(bio)
    xlsx_bytes = bio.getvalue()

    filename = f"Urenoverzicht_{month_first.strftime('%Y-%m')}.xlsx"
    ts = timezone.now().strftime("%Y%m%d_%H%M%S")
    storage_path = f"tmp/uren_overzicht/{ts}_{filename}"
    storage_path = default_storage.save(storage_path, ContentFile(xlsx_bytes))

    return UrenExportResult(
        month=month_first,
        xlsx_storage_path=storage_path,
        filename=filename,
        row_count=len(user_ids),
    )
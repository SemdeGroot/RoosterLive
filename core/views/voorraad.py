# core/views/voorraad.py

import csv
import io
import re
from pathlib import Path

from openpyxl import load_workbook

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, HttpResponse
from django.db import transaction
from django.template.loader import render_to_string
from django.utils import timezone

from ..forms import AvailabilityUploadForm
from ..models import VoorraadItem, Organization
from ._helpers import can
from core.tasks import send_voorraad_html_task


ZI_RE = re.compile(r"^\d{8}$")
TRAILING_DOTZERO_RE = re.compile(r"[,.]0$")


def _clean_cell(value) -> str:
    """
    Normaliseert celwaarden naar string:
    - None -> ""
    - str/int/float -> string + strip
    - verwijdert Excel-artifacten: '12345678.0' of '12345678,0' -> '12345678'
    """
    if value is None:
        return ""
    s = str(value).strip()
    s = TRAILING_DOTZERO_RE.sub("", s)
    return s


def _read_csv_rows(uploaded_file):
    content = uploaded_file.read()
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,|\t,")
    except csv.Error:
        dialect = csv.excel

    f = io.StringIO(text)
    reader = csv.reader(f, dialect)

    rows = []
    for row in reader:
        rows.append([_clean_cell(c) for c in row])
    return rows


def _read_xlsx_rows(uploaded_file):
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([_clean_cell(c) for c in row])
    return rows


@login_required
def medications_view(request):
    # 1. Check Rechten
    if not can(request.user, "can_view_av_medications"):
        return HttpResponseForbidden("Geen toegang.")

    can_edit = can(request, "can_upload_voorraad")
    form = AvailabilityUploadForm()

    # === UPLOAD VERWERKING ===
    if request.method == "POST":
        if not can_edit:
            return HttpResponseForbidden("Geen uploadrechten.")

        form = AvailabilityUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data["file"]
            ext = Path(f.name).suffix.lower()

            # LET OP: openpyxl ondersteunt géén .xls
            if ext not in (".xlsx", ".csv"):
                messages.error(request, "Alleen CSV of .xlsx Excel toegestaan.")
            else:
                try:
                    # --- A. BESTAND INLEZEN ---
                    if ext == ".csv":
                        raw_rows = _read_csv_rows(f)
                    else:
                        raw_rows = _read_xlsx_rows(f)

                    cleaned_rows = []
                    for r in raw_rows:
                        if not r:
                            continue
                        if all(c == "" for c in r):
                            continue
                        cleaned_rows.append(r)

                    if not cleaned_rows:
                        messages.error(request, "Het bestand bevat geen geldige regels.")
                        ctx = {
                            "can_edit": can_edit,
                            "form": form,
                            "title": "Voorraad",
                            "has_data": False,
                            "rows": [],
                            "columns": [],
                            "apotheken": Organization.objects.filter(
                                org_type=Organization.ORG_TYPE_APOTHEEK
                            ).order_by("name"),
                        }
                        return render(request, "voorraad/index.html", ctx)

                    # --- B. HEADER + MIN 2 KOLOMMEN ---
                    header = [h.strip() for h in cleaned_rows[0]]
                    while header and header[-1] == "":
                        header.pop()

                    if len(header) < 2:
                        raise ValueError("Het bestand moet minimaal 2 kolommen hebben (ZI-nummer en Naam).")

                    data_rows = cleaned_rows[1:]

                    filtered = []
                    for r in data_rows:
                        if len(r) < len(header):
                            r = r + [""] * (len(header) - len(r))

                        zi = r[0]
                        naam = r[1]
                        if zi == "" and naam == "":
                            continue
                        filtered.append(r)

                    if not filtered:
                        messages.error(request, "Het bestand bevat geen geldige regels.")
                        ctx = {
                            "can_edit": can_edit,
                            "form": form,
                            "title": "Voorraad",
                            "has_data": False,
                            "rows": [],
                            "columns": [],
                            "apotheken": Organization.objects.filter(
                                org_type=Organization.ORG_TYPE_APOTHEEK
                            ).order_by("name"),
                        }
                        return render(request, "voorraad/index.html", ctx)

                    # --- C. VALIDATIE ---
                    zis = [r[0].strip() for r in filtered]

                    for i, zi in enumerate(zis):
                        if not ZI_RE.match(zi):
                            foute_rij = i + 2
                            messages.error(
                                request,
                                f"Fout op rij {foute_rij}: '{zi}' is geen geldig ZI-nummer (moet 8 cijfers zijn)."
                            )
                            ctx = {
                                "can_edit": can_edit,
                                "form": form,
                                "title": "Voorraad",
                                "has_data": False,
                                "rows": [],
                                "columns": [],
                                "apotheken": Organization.objects.filter(
                                    org_type=Organization.ORG_TYPE_APOTHEEK
                                ).order_by("name"),
                            }
                            return render(request, "voorraad/index.html", ctx)

                    seen = set()
                    duplicate = None
                    for zi in zis:
                        if zi in seen:
                            duplicate = zi
                            break
                        seen.add(zi)

                    if duplicate:
                        messages.error(
                            request,
                            f"Validatiefout: ZI-nummer '{duplicate}' komt meerdere keren voor in het bestand."
                        )
                        ctx = {
                            "can_edit": can_edit,
                            "form": form,
                            "title": "Voorraad",
                            "has_data": False,
                            "rows": [],
                            "columns": [],
                            "apotheken": Organization.objects.filter(
                                org_type=Organization.ORG_TYPE_APOTHEEK
                            ).order_by("name"),
                        }
                        return render(request, "voorraad/index.html", ctx)

                    namen = [r[1] for r in filtered]
                    if not any("paracetamol" in (n or "").lower() for n in namen):
                        messages.warning(
                            request,
                            "Let op: Geen 'Paracetamol' gevonden. Weet je zeker dat de kolommen goed staan?"
                        )

                    # --- D. SYNC LOGICA (VEILIG OPSLAAN) ---
                    file_zi_set = set()
                    existing_items_map = {item.zi_nummer: item for item in VoorraadItem.objects.all()}

                    to_create = []
                    to_update = []

                    meta_columns = header[2:]

                    for r in filtered:
                        zi = r[0].strip()
                        naam = r[1].strip()

                        meta_values = r[2:2 + len(meta_columns)]
                        if len(meta_values) < len(meta_columns):
                            meta_values += [""] * (len(meta_columns) - len(meta_values))

                        metadata = dict(zip(meta_columns, meta_values))

                        file_zi_set.add(zi)

                        if zi in existing_items_map:
                            item = existing_items_map[zi]
                            if item.naam != naam or (item.metadata or {}) != metadata:
                                item.naam = naam
                                item.metadata = metadata
                                to_update.append(item)
                        else:
                            to_create.append(
                                VoorraadItem(
                                    zi_nummer=zi,
                                    naam=naam,
                                    metadata=metadata,
                                )
                            )

                    with transaction.atomic():
                        if to_create:
                            VoorraadItem.objects.bulk_create(to_create)

                        if to_update:
                            VoorraadItem.objects.bulk_update(to_update, ["naam", "metadata"])

                        items_to_delete = VoorraadItem.objects.exclude(zi_nummer__in=file_zi_set)
                        deleted_items_count = items_to_delete.count()
                        items_to_delete.delete()

                    messages.success(
                        request,
                        f"Verwerking gereed: {len(to_create)} toegevoegd, {len(to_update)} geüpdatet, {deleted_items_count} verwijderd."
                    )

                except Exception as e:
                    messages.error(request, f"Fout bij verwerken: {e}")

    # === DATA OPHALEN VOOR TABEL ===
    all_items = VoorraadItem.objects.all()

    rows = []
    columns = []

    if all_items.exists():
        first_item = all_items.first()
        columns = ["ZI Nummer", "Naam"]
        if first_item.metadata:
            columns.extend(first_item.metadata.keys())

        for item in all_items:
            row_data = [item.zi_nummer, item.naam]
            if item.metadata:
                row_data.extend(item.metadata.values())
            rows.append(row_data)

    apotheken = Organization.objects.filter(
        org_type=Organization.ORG_TYPE_APOTHEEK
    ).order_by("name")

    ctx = {
        "can_edit": can_edit,
        "form": form,
        "columns": columns,
        "rows": rows,
        "title": "Voorraad",
        "has_data": len(rows) > 0,
        "apotheken": apotheken,
    }
    return render(request, "voorraad/index.html", ctx)


@login_required
def export_voorraad_html(request):
    if not can(request.user, "can_view_av_medications"):
        return HttpResponseForbidden("Geen toegang.")

    contact_email = "baxterezorg@apotheekjansen.com"
    items = VoorraadItem.objects.all().order_by("naam", "zi_nummer")

    context = {
        "items": items,
        "generated_at": timezone.localtime(timezone.now()),
        "contact_email": contact_email,
        "logo_url": request.build_absolute_uri("/static/img/app_icon-1024x1024.png"),
    }

    html = render_to_string("voorraad/export/voorraad_lijst.html", context, request=request)
    filename = f"Baxtervoorraad_ApoJansen_{timezone.now().strftime('%d-%m-%Y')}.html"

    resp = HttpResponse(html, content_type="text/html; charset=utf-8")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required
def email_voorraad_html(request):
    if not can(request.user, "can_upload_voorraad"):
        return HttpResponseForbidden("Geen toegang.")

    if request.method == "POST":
        org_ids = request.POST.getlist("recipients")
        org_ids = [int(i) for i in org_ids if i.isdigit()]

        if org_ids:
            send_voorraad_html_task.delay(org_ids)
            messages.success(
                request,
                f"De voorraad wordt op de achtergrond verstuurd naar {len(org_ids)} ontvangers."
            )
        else:
            messages.warning(request, "Geen ontvangers geselecteerd.")

    return redirect("medications")

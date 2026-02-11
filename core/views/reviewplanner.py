from datetime import datetime
from dateutil.relativedelta import relativedelta

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponseForbidden, JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.utils import timezone

from core.models import ReviewPlanner, MedicatieReviewAfdeling
from core.forms import ReviewPlannerForm
from ._helpers import can

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

User = get_user_model()


def _eligible_medicatiebeoordeling_users():
    qs = User.objects.filter(is_active=True).order_by("first_name", "last_name", "username")
    return [u for u in qs if can(u, "can_perform_medicatiebeoordeling")]


def _parse_dmy(s: str):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%d-%m-%Y").date()
    except Exception:
        return None


def _parse_hhmm(s: str):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%H:%M").time()
    except Exception:
        return None


def _display_name(u) -> str:
    if not u:
        return ""
    full = (u.get_full_name() or "").strip()
    return full or getattr(u, "username", "") or str(u.pk)


def _serialize_row(obj: ReviewPlanner):
    return {
        "id": obj.id,
        "datum": obj.datum.strftime("%d-%m-%Y") if obj.datum else "",
        "afdeling_id": obj.afdeling_id or "",
        "status": obj.status,
        "status_label": obj.get_status_display(),
        "arts": obj.arts or "",
        "tijd": obj.tijd.strftime("%H:%M") if obj.tijd else "",
        "voorbereid_door_id": obj.voorbereid_door_id or "",
        "uitgevoerd_door_id": obj.uitgevoerd_door_id or "",
        "bijzonderheden": obj.bijzonderheden or "",
    }


def _excel_apply_common_header_style(ws, header_row: int, ncols: int) -> None:
    header_fill = PatternFill("solid", fgColor="E5E7EB")
    header_font = Font(bold=True, color="000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align


def _autosize_columns(ws, min_width: int = 12, max_width: int = 65) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def _afdeling_parts(afd: MedicatieReviewAfdeling | None) -> tuple[str, str, str]:
    if not afd:
        return ("", "", "")
    org_name = getattr(getattr(afd, "organisatie", None), "name", "") or ""
    return (afd.afdeling or "", afd.locatie or "", org_name)


@login_required
def reviewplanner(request):
    if not can(request.user, "can_view_reviewplanner"):
        return HttpResponseForbidden("Geen toegang.")

    can_edit = can(request.user, "can_edit_reviewplanner")

    today = timezone.localdate()
    cutoff = today - relativedelta(weeks=8)

    rows = list(
        ReviewPlanner.objects.filter(datum__gte=cutoff)
        .select_related("afdeling", "afdeling__organisatie", "voorbereid_door", "uitgevoerd_door")
        .order_by("datum", "-updated_at", "-id")
    )

    afdelingen = list(
        MedicatieReviewAfdeling.objects.all()
        .order_by("organisatie__name", "afdeling", "locatie")
    )

    eligible_users = _eligible_medicatiebeoordeling_users()

    if request.method == "POST":
        if not can_edit:
            return JsonResponse({"ok": False, "error": "Geen rechten om te bewerken."}, status=403)

        action = request.POST.get("action", "autosave")

        # ------------------------------------------------------------
        # A) Modal upsert
        # ------------------------------------------------------------
        if action == "modal_upsert":
            rid = (request.POST.get("id") or "").strip()
            datum_s = (request.POST.get("datum") or "").strip()
            afdeling_id = (request.POST.get("afdeling_id") or "").strip()
            status = (request.POST.get("status") or ReviewPlanner.STATUS_PREP).strip()
            arts = (request.POST.get("arts") or "").strip()
            tijd_s = (request.POST.get("tijd") or "").strip()

            voorbereid_s = (request.POST.get("voorbereid_door") or "").strip()
            uitgevoerd_s = (request.POST.get("uitgevoerd_door") or "").strip()
            bijz = (request.POST.get("bijzonderheden") or "").strip()

            d = _parse_dmy(datum_s)
            if not d:
                return JsonResponse({"ok": False, "error": "Datum is verplicht (dd-mm-jjjj)."}, status=400)

            if d < today:
                return JsonResponse({"ok": False, "error": "Datum mag niet in het verleden liggen."}, status=400)
            if d < cutoff:
                return JsonResponse({"ok": False, "error": "Datum mag niet verder dan 8 weken terug liggen."}, status=400)

            t = _parse_hhmm(tijd_s)
            if tijd_s and not t:
                return JsonResponse({"ok": False, "error": "Ongeldige tijd."}, status=400)

            afd = None
            if afdeling_id:
                try:
                    afd = MedicatieReviewAfdeling.objects.get(id=int(afdeling_id))
                except Exception:
                    return JsonResponse({"ok": False, "error": "Ongeldige afdeling."}, status=400)

            if status not in dict(ReviewPlanner.STATUS_CHOICES):
                return JsonResponse({"ok": False, "error": "Ongeldige status."}, status=400)

            voorbereid_user = None
            if voorbereid_s:
                try:
                    voorbereid_user = User.objects.get(id=int(voorbereid_s), is_active=True)
                except Exception:
                    return JsonResponse({"ok": False, "error": "Ongeldige voorbereid door."}, status=400)

            uitgevoerd_user = None
            if uitgevoerd_s:
                try:
                    uitgevoerd_user = User.objects.get(id=int(uitgevoerd_s), is_active=True)
                except Exception:
                    return JsonResponse({"ok": False, "error": "Ongeldige uitgevoerd door."}, status=400)

            if voorbereid_user and not can(voorbereid_user, "can_perform_medicatiebeoordeling"):
                return JsonResponse({"ok": False, "error": "Voorbereid door heeft geen juiste rol."}, status=400)
            if uitgevoerd_user and not can(uitgevoerd_user, "can_perform_medicatiebeoordeling"):
                return JsonResponse({"ok": False, "error": "Uitgevoerd door heeft geen juiste rol."}, status=400)

            with transaction.atomic():
                if rid:
                    obj = ReviewPlanner.objects.select_for_update().get(id=int(rid))
                else:
                    obj = ReviewPlanner(created_by=request.user)

                obj.datum = d
                obj.afdeling = afd
                obj.status = status
                obj.arts = arts
                obj.tijd = t
                obj.voorbereid_door = voorbereid_user
                obj.uitgevoerd_door = uitgevoerd_user
                obj.bijzonderheden = bijz
                obj.updated_by = request.user
                obj.save()

            obj = ReviewPlanner.objects.select_related("voorbereid_door", "uitgevoerd_door").get(id=obj.id)
            return JsonResponse({"ok": True, "row": _serialize_row(obj)})

        # ------------------------------------------------------------
        # B) Autosave – bulk rows
        # ------------------------------------------------------------
        if action == "autosave":
            ids = request.POST.getlist("row_id")
            datums = request.POST.getlist("row_datum")
            afds = request.POST.getlist("row_afdeling")
            statuses = request.POST.getlist("row_status")
            arts_list = request.POST.getlist("row_arts")
            tijden = request.POST.getlist("row_tijd")
            voorbereid_list = request.POST.getlist("row_voorbereid_door")
            uitgevoerd_list = request.POST.getlist("row_uitgevoerd_door")
            bijz_list = request.POST.getlist("row_bijzonderheden")

            n = min(
                len(ids), len(datums), len(afds), len(statuses),
                len(arts_list), len(tijden),
                len(voorbereid_list), len(uitgevoerd_list),
                len(bijz_list),
            )

            try:
                with transaction.atomic():
                    for i in range(n):
                        rid = (ids[i] or "").strip()
                        datum_s = (datums[i] or "").strip()
                        afd_s = (afds[i] or "").strip()
                        status = (statuses[i] or ReviewPlanner.STATUS_PREP).strip()
                        arts = (arts_list[i] or "").strip()
                        tijd_s = (tijden[i] or "").strip()
                        voorbereid_s = (voorbereid_list[i] or "").strip()
                        uitgevoerd_s = (uitgevoerd_list[i] or "").strip()
                        bijz = (bijz_list[i] or "").strip()

                        all_empty = (
                            not datum_s and not afd_s and not arts and not tijd_s
                            and not voorbereid_s and not uitgevoerd_s and not bijz
                        )
                        if rid and all_empty:
                            ReviewPlanner.objects.filter(id=int(rid)).delete()
                            continue
                        if (not rid) and all_empty:
                            continue

                        d = _parse_dmy(datum_s)
                        if datum_s and not d:
                            return JsonResponse({"ok": False, "error": "Ongeldige datum."}, status=400)

                        if d and d < today:
                            return JsonResponse({"ok": False, "error": "Datum mag niet in het verleden liggen."}, status=400)
                        if d and d < cutoff:
                            return JsonResponse({"ok": False, "error": "Datum mag niet verder dan 8 weken terug liggen."}, status=400)

                        t = _parse_hhmm(tijd_s)
                        if tijd_s and not t:
                            return JsonResponse({"ok": False, "error": "Ongeldige tijd."}, status=400)

                        afd = None
                        if afd_s:
                            try:
                                afd = MedicatieReviewAfdeling.objects.get(id=int(afd_s))
                            except Exception:
                                return JsonResponse({"ok": False, "error": "Ongeldige afdeling."}, status=400)

                        if status not in dict(ReviewPlanner.STATUS_CHOICES):
                            return JsonResponse({"ok": False, "error": "Ongeldige status."}, status=400)

                        voorbereid_user = None
                        if voorbereid_s:
                            try:
                                voorbereid_user = User.objects.get(id=int(voorbereid_s), is_active=True)
                            except Exception:
                                return JsonResponse({"ok": False, "error": "Ongeldige voorbereid door."}, status=400)

                        uitgevoerd_user = None
                        if uitgevoerd_s:
                            try:
                                uitgevoerd_user = User.objects.get(id=int(uitgevoerd_s), is_active=True)
                            except Exception:
                                return JsonResponse({"ok": False, "error": "Ongeldige uitgevoerd door."}, status=400)

                        if voorbereid_user and not can(voorbereid_user, "can_perform_medicatiebeoordeling"):
                            return JsonResponse({"ok": False, "error": "Voorbereid door heeft geen juiste rol."}, status=400)
                        if uitgevoerd_user and not can(uitgevoerd_user, "can_perform_medicatiebeoordeling"):
                            return JsonResponse({"ok": False, "error": "Uitgevoerd door heeft geen juiste rol."}, status=400)

                        if rid:
                            obj = ReviewPlanner.objects.select_for_update().get(id=int(rid))
                        else:
                            obj = ReviewPlanner(created_by=request.user)

                        obj.datum = d
                        obj.afdeling = afd
                        obj.status = status
                        obj.arts = arts
                        obj.tijd = t
                        obj.voorbereid_door = voorbereid_user
                        obj.uitgevoerd_door = uitgevoerd_user
                        obj.bijzonderheden = bijz
                        obj.updated_by = request.user
                        obj.save()

                return JsonResponse({"ok": True})
            except Exception:
                return JsonResponse({"ok": False, "error": "Opslaan mislukt."}, status=400)

        return redirect("reviewplanner")

    context = {
        "rows": rows,
        "afdelingen": afdelingen,
        "can_edit": can_edit,
        "status_choices": ReviewPlanner.STATUS_CHOICES,
        "eligible_users": eligible_users,
        "cutoff_iso": cutoff.isoformat(),
        "today_iso": today.isoformat(),
    }
    return render(request, "reviewplanner/index.html", context)


@login_required
def reviewplanner_export_overview(request):
    if not can(request.user, "can_view_reviewplanner"):
        return HttpResponseForbidden("Geen toegang.")

    if not can(request.user, "can_edit_reviewplanner"):
        return HttpResponseForbidden("Geen rechten om te exporteren.")

    qs = (
        ReviewPlanner.objects.select_related(
            "afdeling", "afdeling__organisatie",
            "voorbereid_door", "uitgevoerd_door",
        )
        .all()
        .order_by("datum", "id")
    )

    rows_by_year: dict[str, list[ReviewPlanner]] = {}
    no_date: list[ReviewPlanner] = []

    for r in qs:
        if not r.datum:
            no_date.append(r)
            continue
        y = str(r.datum.year)
        rows_by_year.setdefault(y, []).append(r)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    # Bijzonderheden als laatste
    headers = [
        "Datum", "Afdeling", "Locatie", "Organisatie",
        "Status", "Arts", "Tijd",
        "Voorbereid door", "Uitgevoerd door",
        "Bijzonderheden",
    ]

    def add_sheet(title: str, sheet_rows: list[ReviewPlanner], with_date: bool):
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        _excel_apply_common_header_style(ws, header_row=1, ncols=len(headers))

        for r in sheet_rows:
            afd_name, loc, org = _afdeling_parts(r.afdeling)
            ws.append(
                [
                    (r.datum.strftime("%d-%m-%Y") if (with_date and r.datum) else ""),
                    afd_name,
                    loc,
                    org,
                    r.get_status_display(),
                    r.arts or "",
                    r.tijd.strftime("%H:%M") if r.tijd else "",
                    _display_name(r.voorbereid_door),
                    _display_name(r.uitgevoerd_door),
                    r.bijzonderheden or "",
                ]
            )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:J{ws.max_row}"

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row[0].alignment = Alignment(horizontal="center", vertical="center")  # Datum
            row[1].alignment = Alignment(horizontal="left", vertical="center")    # Afdeling
            row[2].alignment = Alignment(horizontal="left", vertical="center")    # Locatie
            row[3].alignment = Alignment(horizontal="left", vertical="center")    # Org
            row[4].alignment = Alignment(horizontal="center", vertical="center")  # Status
            row[5].alignment = Alignment(horizontal="left", vertical="center")    # Arts
            row[6].alignment = Alignment(horizontal="center", vertical="center")  # Tijd
            row[7].alignment = Alignment(horizontal="left", vertical="center")    # Voorbereid
            row[8].alignment = Alignment(horizontal="left", vertical="center")    # Uitgevoerd
            row[9].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)  # Bijz

        _autosize_columns(ws, min_width=12, max_width=65)

    for year in sorted(rows_by_year.keys()):
        add_sheet(year, rows_by_year[year], with_date=True)

    if no_date:
        add_sheet("Zonder datum", no_date, with_date=False)

    if not wb.sheetnames:
        wb.create_sheet(title=str(timezone.localdate().year))

    bio = BytesIO()
    wb.save(bio)

    stamp = timezone.localdate().strftime("%Y-%m-%d")
    filename = f"Reviewplanner_overzicht_{stamp}.xlsx"

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp